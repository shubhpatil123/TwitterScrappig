import traceback
from tweepy import API 
from tweepy import Cursor
from tweepy.streaming import StreamListener
from tweepy import OAuthHandler
from tweepy import Stream

from textblob import TextBlob
 
import twitter_credentials

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import re
import sys
import csv
import pyttsx3
import openpyxl

# # # # TWITTER CLIENT # # # #
class TwitterClient():
    def __init__(self, twitter_user=None):
        self.auth = TwitterAuthenticator().authenticate_twitter_app()
        self.twitter_client = API(self.auth)

        self.twitter_user = twitter_user

    def get_twitter_client_api(self):
        return self.twitter_client

    def get_user_timeline_tweets(self, num_tweets):
        tweets = []
        for tweet in Cursor(self.twitter_client.user_timeline, id=self.twitter_user).items(num_tweets):
            tweets.append(tweet)
        return tweets

    def get_friend_list(self, num_friends):
        friend_list = []
        for friend in Cursor(self.twitter_client.friends, id=self.twitter_user).items(num_friends):
            friend_list.append(friend)
        return friend_list

    def get_home_timeline_tweets(self, num_tweets):
        home_timeline_tweets = []
        for tweet in Cursor(self.twitter_client.home_timeline, id=self.twitter_user).items(num_tweets):
            home_timeline_tweets.append(tweet)
        return home_timeline_tweets


# # # # TWITTER AUTHENTICATER # # # #
class TwitterAuthenticator():

    def authenticate_twitter_app(self):
        auth = OAuthHandler(twitter_credentials.CONSUMER_KEY, twitter_credentials.CONSUMER_SECRET)
        auth.set_access_token(twitter_credentials.ACCESS_TOKEN, twitter_credentials.ACCESS_TOKEN_SECRET)
        return auth

# # # # TWITTER STREAMER # # # #
class TwitterStreamer():
    """
    Class for streaming and processing live tweets.
    """
    def __init__(self):
        self.twitter_autenticator = TwitterAuthenticator()    

    def stream_tweets(self, fetched_tweets_filename, hash_tag_list):
        # This handles Twitter authetification and the connection to Twitter Streaming API
        listener = TwitterListener(fetched_tweets_filename)
        auth = self.twitter_autenticator.authenticate_twitter_app() 
        stream = Stream(auth, listener)

        # This line filter Twitter Streams to capture data by the keywords: 
        stream.filter(track=hash_tag_list)


# # # # TWITTER STREAM LISTENER # # # #
class TwitterListener(StreamListener):
    """
    This is a basic listener that just prints received tweets to stdout.
    """
    def __init__(self, fetched_tweets_filename):
        self.fetched_tweets_filename = fetched_tweets_filename

    def on_data(self, data):
        try:
            print(data)
            with open(self.fetched_tweets_filename, 'a') as tf:
                tf.write(data)
            return True
        except BaseException as e:
            print("Error on_data %s" % str(e))
        return True
          
    def on_error(self, status):
        if status == 420:
            # Returning False on_data method in case rate limit occurs.
            return False
        print(status)


class TweetAnalyzer():
    """
    Functionality for analyzing and categorizing content from tweets.
    """

    def clean_tweet(self, tweet):
        return ' '.join(re.sub("(@[A-Za-z0-9]+)|([^0-9A-Za-z \t])|(\w+:\/\/\S+)", " ", str(tweet)).split())

    

    def tweets_to_data_frame(self, tweets):
        df={}
        df = pd.DataFrame(data=[tweet.full_text for tweet in tweets], columns=['tweets'])

        df['id'] = [tweet.id for tweet in tweets]
        df['len'] = [len(tweet.full_text) for tweet in tweets]
        df['date'] = [tweet.created_at for tweet in tweets]
        df['source'] = [tweet.source for tweet in tweets]
        df['likes'] = [tweet.favorite_count for tweet in tweets]
        df['retweets'] = [tweet.retweet_count for tweet in tweets]
        print(df)
        return df

    ####function for sentimental analysis
    def get_tweet_sentiment(self, tweet): 
        ''' 
        Utility function to classify sentiment of passed tweet 
        using textblob's sentiment method 
        '''
        # create TextBlob object of passed tweet text 
        analysis = TextBlob(tweet)
        print(self.clean_tweet(tweet),analysis.sentiment.polarity)
        # set sentiment 
        if analysis.sentiment.polarity > 0: 
            return 'positive'
        elif analysis.sentiment.polarity == 0: 
            return 'neutral'
        else: 
            return 'negative'

    def get_tweets(self, tweets):
        print(len(tweets))
        tmp=[]
        ''' 
        Main function to parse tweets. 
        '''
        #fethched tweets
        tweets_fethched = [] 
  
        try: 
            tweets_fethched=tweets 
  
            # parsing tweets one by one 
            for tweet in tweets:
                print('***************')
                # empty dictionary to store required params of a tweet 
                parsed_tweet = {} 
  
                # saving text of tweet 
                parsed_tweet['text'] = tweet 
                # saving sentiment of tweet 
                parsed_tweet['sentiment'] = self.get_tweet_sentiment(tweet.full_text)
                
  
                # appending parsed tweet to tweets list 
                if tweet.retweet_count > 0: 
                    # if tweet has retweets, ensure that it is appended only once 
                    if parsed_tweet not in tweets: 
                        tmp.append(parsed_tweet) 
                else: 
                    tmp.append(parsed_tweet) 
  
            # return parsed tweets 
            return tmp 
  
        except: 
            # print error (if any) 
            print("Error : ",traceback.print_exc() ) 

 
if __name__ == '__main__':

    twitter_client = TwitterClient()
    tweet_analyzer = TweetAnalyzer()
    tweets_for_csv = []
    api = twitter_client.get_twitter_client_api()

    tweets = api.user_timeline(screen_name="narendramodi", count=100,tweet_mode="extended")
    lines=['ID','LEN','DATE','SOURCE','LIKES','RETWEETS']
    df = tweet_analyzer.tweets_to_data_frame(tweets)
    
    #write to a new csv file from the array of tweets
    #print(df)
    #outfile = 'pune' + "_tweets.csv"
    print(tweets)
    #with open(outfile, 'w') as writeFile:
    #   writer = csv.writer(writeFile)
    #    writer.writerow(lines)
    # Call a Workbook() function of openpyxl  
    # to create a new blank Workbook object 
    wb = openpyxl.Workbook() 


    
    # Get workbook active sheet   
    # from the active attribute.  
    sheet = wb.active
    sheet.cell(row = 1, column = 1).value = ' ID '
    sheet.cell(row = 1, column = 2).value = ' LEN '
    sheet.cell(row = 1, column = 3).value = ' DATE '
    sheet.cell(row = 1, column = 4).value = ' SOURCE '
    sheet.cell(row = 1, column = 5).value = ' LIKES '
    sheet.cell(row = 1, column = 6).value = ' RETWEETS '
    



    
    tmp = []
    tweets_for_csv = [tweet.full_text for tweet in tweets] # CSV file created  
    for j in tweets_for_csv: 
    #Appending tweets to the empty array tmp 
        tmp.append(j)  
  
    # Printing the tweets
    print('***************SP *****************')
    

  
#    print("writing to " + outfile)
    for i in range(len(df['id'])):
        sheet.cell(row = i+2, column = 1).value =df['id'][i]
        sheet.cell(row = i+2, column = 2).value =df['len'][i]
        sheet.cell(row = i+2, column = 3).value =df['date'][i]
        sheet.cell(row = i+2, column = 4).value =df['source'][i]
        sheet.cell(row = i+2, column = 5).value =df['likes'][i]
        sheet.cell(row = i+2, column = 6).value =df['retweets'][i]
        
    # save the file 
    wb.save('modi_tweets.xlsx')

    print('--------------new edit here -----------------')
    print(tmp)
    tmp=tweet_analyzer.get_tweets(tweets)
    ptweets = [tweet for tweet in tmp if tweet['sentiment'] == 'positive'] 
    # percentage of positive tweets 
    print("Positive tweets percentage: {} %".format(100*len(ptweets)/len(tmp))) 
    # picking negative tweets from tweets 
    ntweets = [tweet for tweet in tmp if tweet['sentiment'] == 'negative'] 
    # percentage of negative tweets 
    print("Negative tweets percentage: {} %".format(100*len(ntweets)/len(tmp))) 
    # percentage of neutral tweets 
    #print("Neutral tweets percentage: {} %".format(100*len(tweets\p - ntweets - ptweets)/len(tweets))) 
  
    # printing first 5 positive tweets 
    print("\n\nPositive tweets:") 
    for tweet in ptweets[:10]: 
        print(tweet['text']) 
  
    # printing first 5 negative tweets 
    print("\n\nNegative tweets:") 
    for tweet in ntweets[:10]: 
        print(tweet['text']) 







    
    #with open(outfile, 'w') as file:
        
        #for i in range(len(df['id'])):
            #QW=[]
            #writer = csv.writer(file, delimiter=',')
            #QW=list([df['id'][i],df['len'][i],df['date'][i],df['source'][i],df['likes'][i],df['retweets'][i]])
            #print(QW)
            #writer.writerow(map(lambda x: [x], QW))
    
    #Twitter Trends function
    trends1 = api.trends_place(1) # from the end of your code
# trends1 is a list with only one element in it, which is a 
# dict which we'll put in data.
    data = trends1[0] 
# grab the trends
    trends = data['trends']
# grab the name from each trend
    names = [trend['name'] for trend in trends]
# put all the names together with a ' ' separating them
    trendsName = ' '.join(names)
    print(trendsName)
    for a in names:
        print(a)
            

            
